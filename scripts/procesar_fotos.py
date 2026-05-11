"""Procesa el volcado de fotos del móvil para un encargo.

Origen: carpeta de Google Drive del encargo (Drive for Desktop), donde Javier
sube las fotos directamente desde la galería del móvil. Esto preserva los
metadatos EXIF (DateTimeOriginal), a diferencia del flujo antiguo de WhatsApp.

Fase 1 — copiado + renombrado:
    Copia las fotos desde --source (Drive) a 04_inspeccion/fotos/.
    Lee EXIF DateTimeOriginal de cada una (con fallback a mtime si falta).
    Ordena por fecha real y renombra a P001.jpg, P002.jpg…
    Genera 04_inspeccion/fotos/INDICE.md.
    Genera 04_inspeccion/clasificacion_fotos.xlsx con MINIATURAS embebidas
    para que Javier rellene Deficiencia + Descripción de cada foto.
    Si todas las fotos tienen EXIF, escribe 04_inspeccion/fecha_inspeccion.txt
    con el rango de fechas detectado.

    --source es opcional: si se omite, opera sólo sobre lo que ya hay en
    04_inspeccion/fotos/ (compatible con flujo antiguo / volcados manuales).

Fase 2 — aplicar Excel (RECOMENDADO):
    Lee 04_inspeccion/clasificacion_fotos.xlsx (rellenado por Javier),
    mueve cada foto a 04_inspeccion/fotos_renombradas/ con nombre
    P001_DEFNN_slug.jpg (slug derivado de la descripción), y escribe
    _skill_workspace/fotos_descripciones.json con las descripciones
    para que redactar_v1.py las use como caption.

Fase 2 — clasificación por rangos (LEGACY):
    python procesar_fotos.py --encargo "<encargo>" --clasificar \\
        --map "1-12:DEF01:fachada_norte"

Uso:
    # Fase 1 — desde carpeta de Drive
    python procesar_fotos.py --encargo "<encargo>" --renombrar \\
        --source "G:\\Mi unidad\\Encargos\\2026_005\\fotos"

    # Fase 2 — aplicar Excel
    python procesar_fotos.py --encargo "<encargo>" --aplicar-excel
"""
from __future__ import annotations

import argparse
import datetime as dt
import json
import os
import re
import shutil
import sys
import unicodedata
from pathlib import Path

try:
    sys.stdout.reconfigure(encoding="utf-8")
    sys.stderr.reconfigure(encoding="utf-8")
except Exception:
    pass

EXT_VALIDAS = {".jpg", ".jpeg", ".png", ".heic", ".heif", ".webp"}

try:
    from PIL import Image, ExifTags
    _PIL_OK = True
except Exception:
    _PIL_OK = False

try:
    import pillow_heif
    pillow_heif.register_heif_opener()
    _HEIF_OK = True
except Exception:
    _HEIF_OK = False


def _exif_datetime(p: Path) -> dt.datetime | None:
    """Lee EXIF DateTimeOriginal (o DateTime) de la foto. None si no se puede."""
    if not _PIL_OK:
        return None
    if p.suffix.lower() in {".heic", ".heif"} and not _HEIF_OK:
        return None
    try:
        with Image.open(p) as img:
            exif = img.getexif()
            if not exif:
                return None
            tag_map = {ExifTags.TAGS.get(k, k): v for k, v in exif.items()}
            ifd = exif.get_ifd(0x8769) if hasattr(exif, "get_ifd") else {}
            ifd_map = {ExifTags.TAGS.get(k, k): v for k, v in ifd.items()} if ifd else {}
            for key in ("DateTimeOriginal", "DateTimeDigitized", "DateTime"):
                val = ifd_map.get(key) or tag_map.get(key)
                if val:
                    s = str(val).strip()
                    for fmt in ("%Y:%m:%d %H:%M:%S", "%Y-%m-%d %H:%M:%S"):
                        try:
                            return dt.datetime.strptime(s, fmt)
                        except ValueError:
                            continue
    except Exception:
        return None
    return None


def fecha_orden(p: Path) -> tuple[dt.datetime, str]:
    """Devuelve (timestamp, fuente) para ordenar. fuente in {'exif','mtime'}."""
    e = _exif_datetime(p)
    if e is not None:
        return e, "exif"
    return dt.datetime.fromtimestamp(p.stat().st_mtime), "mtime"


def listar_fotos(carpeta: Path) -> list[Path]:
    if not carpeta.exists():
        return []
    fotos = [p for p in carpeta.iterdir() if p.is_file() and p.suffix.lower() in EXT_VALIDAS]
    fotos.sort(key=lambda p: (fecha_orden(p)[0], p.name))
    return fotos


def _slugify(s: str, max_len: int = 40) -> str:
    """Convierte texto libre en slug snake_case sin acentos."""
    s = unicodedata.normalize("NFKD", s).encode("ascii", "ignore").decode()
    s = re.sub(r"[^a-zA-Z0-9]+", "_", s.strip().lower())
    s = re.sub(r"_+", "_", s).strip("_")
    return s[:max_len] or "foto"


def agrupar_por_tiempo(fotos: list[Path], gap_minutos: int = 5) -> list[int]:
    """Devuelve lista paralela a `fotos` con el id de grupo (1, 2, 3...)."""
    if not fotos:
        return []
    grupos = [1]
    prev_ts = fecha_orden(fotos[0])[0]
    grupo_actual = 1
    for foto in fotos[1:]:
        ts = fecha_orden(foto)[0]
        delta = (ts - prev_ts).total_seconds() / 60
        if delta > gap_minutos:
            grupo_actual += 1
        grupos.append(grupo_actual)
        prev_ts = ts
    return grupos


def _miniatura_path(src: Path, workspace: Path, max_side: int = 200) -> Path:
    """Genera miniatura JPG en workspace/thumbs/. Devuelve ruta. Convierte HEIC."""
    out_dir = workspace / "thumbs"
    out_dir.mkdir(parents=True, exist_ok=True)
    out = out_dir / (src.stem + ".jpg")
    if out.exists() and out.stat().st_mtime >= src.stat().st_mtime:
        return out
    if not _PIL_OK:
        return src
    try:
        with Image.open(src) as img:
            img = img.convert("RGB")
            img.thumbnail((max_side, max_side))
            img.save(out, "JPEG", quality=80)
        return out
    except Exception as e:
        print(f"WARN: no se pudo generar miniatura de {src.name}: {e}", file=sys.stderr)
        return src


def _leer_slugs_deficiencias(encargo: Path) -> list[str]:
    """Lee deficiencias.md y devuelve lista de slugs."""
    p = encargo / "deficiencias.md"
    if not p.exists():
        return []
    slugs = []
    for line in p.read_text(encoding="utf-8").splitlines():
        m = re.match(r"^\s*\d+\.\s+\*\*([a-z0-9_]+)\*\*", line)
        if m:
            slugs.append(m.group(1))
    return slugs


def generar_excel_clasificacion(encargo: Path) -> Path:
    """Genera 04_inspeccion/clasificacion_fotos.xlsx con miniaturas + columnas a rellenar."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        from openpyxl.drawing.image import Image as XLImage
        from openpyxl.worksheet.datavalidation import DataValidation
    except ImportError:
        print("ERROR: openpyxl no instalado. pip install openpyxl", file=sys.stderr)
        return Path()

    fotos_dir = encargo / "04_inspeccion" / "fotos"
    workspace = encargo / "_skill_workspace"
    workspace.mkdir(parents=True, exist_ok=True)
    fotos = listar_fotos(fotos_dir)
    if not fotos:
        return Path()

    grupos = agrupar_por_tiempo(fotos)
    slugs = _leer_slugs_deficiencias(encargo)

    wb = Workbook()
    ws = wb.active
    ws.title = "Clasificación fotos"

    headers = ["Foto", "Archivo", "Fecha EXIF", "Grupo", "Miniatura", "Deficiencia", "Descripción", "Notas"]
    bold = Font(bold=True, color="FFFFFF")
    fill_hdr = PatternFill("solid", fgColor="305496")
    centered = Alignment(horizontal="center", vertical="center", wrap_text=True)
    wrap_left = Alignment(horizontal="left", vertical="center", wrap_text=True)
    thin = Side(border_style="thin", color="999999")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for col, h in enumerate(headers, start=1):
        c = ws.cell(row=1, column=col, value=h)
        c.font = bold
        c.fill = fill_hdr
        c.alignment = centered
        c.border = border

    widths = {"A": 8, "B": 28, "C": 18, "D": 8, "E": 26, "F": 22, "G": 50, "H": 25}
    for col_letter, w in widths.items():
        ws.column_dimensions[col_letter].width = w
    ws.row_dimensions[1].height = 28

    # Validación de datos para la columna Deficiencia
    if slugs:
        opciones = ",".join(slugs + ["general", "descartar"])
        dv = DataValidation(type="list", formula1=f'"{opciones}"', allow_blank=True)
        dv.error = "Usa uno de los slugs de deficiencias.md, 'general' o 'descartar'"
        dv.errorTitle = "Valor no válido"
        dv.prompt = "Selecciona deficiencia, 'general' o 'descartar'"
        dv.promptTitle = "Deficiencia"
        ws.add_data_validation(dv)

    for i, foto in enumerate(fotos, start=2):
        ts, fuente = fecha_orden(foto)
        m = re.match(r"^P(\d+)", foto.stem)
        p_id = m.group(0) if m else foto.stem
        ws.cell(row=i, column=1, value=p_id).alignment = centered
        ws.cell(row=i, column=2, value=foto.name).alignment = wrap_left
        ws.cell(row=i, column=3, value=ts.strftime("%Y-%m-%d %H:%M")).alignment = centered
        ws.cell(row=i, column=4, value=f"G{grupos[i-2]}").alignment = centered
        # Miniatura
        thumb = _miniatura_path(foto, workspace)
        try:
            xlimg = XLImage(str(thumb))
            xlimg.width = 160
            xlimg.height = int(160 * xlimg.height / xlimg.width) if xlimg.width else 120
            ws.add_image(xlimg, f"E{i}")
        except Exception as e:
            ws.cell(row=i, column=5, value=f"(no se pudo embeber: {e})")
        ws.row_dimensions[i].height = 130
        for col in range(1, len(headers) + 1):
            ws.cell(row=i, column=col).border = border
        ws.cell(row=i, column=6).alignment = centered
        ws.cell(row=i, column=7).alignment = wrap_left
        ws.cell(row=i, column=8).alignment = wrap_left
        if slugs:
            dv.add(f"F{i}")

    # Hoja "Instrucciones"
    ws2 = wb.create_sheet("Instrucciones")
    instrucciones = [
        "INSTRUCCIONES",
        "",
        "1. En la pestaña 'Clasificación fotos', rellena DEFICIENCIA y DESCRIPCIÓN de cada foto.",
        "2. DEFICIENCIA debe coincidir con un slug de deficiencias.md, o ser 'general' o 'descartar'.",
        "   - 'general': foto de contexto (no de un defecto concreto). Se incluirá en antecedentes o portada.",
        "   - 'descartar': foto que NO se quiere en el informe (mal enfoque, repetida, etc.).",
        "3. DESCRIPCIÓN: texto libre breve (5-15 palabras) sobre qué se ve. Será el pie de imagen en el informe.",
        "   Ej: 'humedad en muro este garaje con eflorescencias salinas'",
        "4. Notas: opcional, observaciones para ti mismo.",
        "5. El GRUPO es sólo una pista: fotos seguidas en el tiempo (gap >5 min entre fotos = nuevo grupo).",
        "6. Cuando termines, guarda y dime 'aplica la clasificación' (o ejecuta procesar_fotos.py --aplicar-excel).",
        "",
        "SLUGS DISPONIBLES (de deficiencias.md):",
    ]
    for line in instrucciones:
        ws2.append([line])
    for slug in slugs:
        ws2.append([f"  - {slug}"])
    ws2.append(["  - general"])
    ws2.append(["  - descartar"])
    ws2.column_dimensions["A"].width = 90

    xlsx_path = encargo / "04_inspeccion" / "clasificacion_fotos.xlsx"
    wb.save(str(xlsx_path))
    return xlsx_path


def fase_aplicar_excel(encargo: Path) -> int:
    """Lee clasificacion_fotos.xlsx, mueve fotos a fotos_renombradas/ y genera JSON."""
    try:
        from openpyxl import load_workbook
    except ImportError:
        print("ERROR: openpyxl no instalado.", file=sys.stderr)
        return 1

    xlsx = encargo / "04_inspeccion" / "clasificacion_fotos.xlsx"
    if not xlsx.exists():
        print(f"ERROR: no existe {xlsx}. Ejecuta --renombrar primero.", file=sys.stderr)
        return 1

    fotos_dir = encargo / "04_inspeccion" / "fotos"
    dest_dir = encargo / "04_inspeccion" / "fotos_renombradas"
    dest_dir.mkdir(parents=True, exist_ok=True)
    workspace = encargo / "_skill_workspace"
    workspace.mkdir(parents=True, exist_ok=True)

    wb = load_workbook(str(xlsx), data_only=True)
    ws = wb["Clasificación fotos"]

    slugs = _leer_slugs_deficiencias(encargo)
    slug_to_def = {slug: i + 1 for i, slug in enumerate(slugs)}

    descripciones: dict[str, dict] = {}
    movidas = 0
    descartadas = 0
    sin_clasificar = []

    for row in ws.iter_rows(min_row=2, values_only=False):
        p_id_cell, archivo_cell, fecha_cell, grupo_cell, _img_cell, def_cell, desc_cell, notas_cell = row[:8]
        archivo = archivo_cell.value
        if not archivo:
            continue
        deficiencia = (def_cell.value or "").strip()
        descripcion = (desc_cell.value or "").strip()
        notas = (notas_cell.value or "").strip()
        p_id = (p_id_cell.value or "").strip()

        if not deficiencia:
            sin_clasificar.append(archivo)
            continue
        if deficiencia.lower() == "descartar":
            descartadas += 1
            descripciones[p_id] = {"archivo": archivo, "deficiencia": "descartar", "descripcion": descripcion, "notas": notas}
            continue

        src = fotos_dir / archivo
        if not src.exists():
            print(f"WARN: no existe {src}", file=sys.stderr)
            continue

        if deficiencia == "general":
            defcod = "GEN"
        elif deficiencia in slug_to_def:
            defcod = f"DEF{slug_to_def[deficiencia]:02d}"
        else:
            print(f"WARN: deficiencia '{deficiencia}' no está en deficiencias.md — se usa GEN", file=sys.stderr)
            defcod = "GEN"

        slug_desc = _slugify(descripcion) if descripcion else _slugify(deficiencia)
        dst = dest_dir / f"{p_id}_{defcod}_{slug_desc}{src.suffix.lower()}"
        if dst.exists() and src.resolve() != dst.resolve():
            dst.unlink()
        shutil.move(str(src), str(dst))
        movidas += 1
        descripciones[p_id] = {
            "archivo_original": archivo,
            "archivo_final": dst.name,
            "deficiencia": deficiencia,
            "defcod": defcod,
            "descripcion": descripcion,
            "notas": notas,
        }

    json_path = workspace / "fotos_descripciones.json"
    json_path.write_text(json.dumps(descripciones, indent=2, ensure_ascii=False), encoding="utf-8")

    print(f"OK — {movidas} fotos clasificadas, {descartadas} descartadas.")
    if sin_clasificar:
        print(f"AVISO: {len(sin_clasificar)} fotos sin Deficiencia rellena (se quedan en fotos/).")
        for a in sin_clasificar[:10]:
            print(f"   - {a}")
        if len(sin_clasificar) > 10:
            print(f"   ... y {len(sin_clasificar)-10} más")
    print(f"Descripciones guardadas en: {json_path}")
    return 0


def copiar_desde_source(source: Path, dest: Path) -> tuple[int, int]:
    """Copia recursivamente fotos de source → dest. Devuelve (copiadas, omitidas)."""
    if not source.exists():
        print(f"ERROR: --source no existe: {source}", file=sys.stderr)
        sys.exit(1)
    dest.mkdir(parents=True, exist_ok=True)
    copiadas = 0
    omitidas = 0
    for src in source.rglob("*"):
        if not src.is_file() or src.suffix.lower() not in EXT_VALIDAS:
            continue
        dst = dest / src.name
        if dst.exists() and dst.stat().st_size == src.stat().st_size:
            omitidas += 1
            continue
        shutil.copy2(str(src), str(dst))
        copiadas += 1
    return copiadas, omitidas


def fase_renombrar(encargo: Path, source: Path | None) -> int:
    fotos_dir = encargo / "04_inspeccion" / "fotos"
    fotos_dir.mkdir(parents=True, exist_ok=True)

    if source is not None:
        c, o = copiar_desde_source(source, fotos_dir)
        print(f"Copiadas {c} fotos desde Drive ({o} ya estaban).")

    fotos = listar_fotos(fotos_dir)
    if not fotos:
        print(f"ERROR: no hay fotos en {fotos_dir}", file=sys.stderr)
        return 1

    # Saltar las que ya están renombradas con el patrón Pxxx
    ya_ok = re.compile(r"^P\d{3,4}(?:_|\.)")
    pendientes = [p for p in fotos if not ya_ok.match(p.name)]
    if not pendientes:
        print(f"Todas las fotos ya están renombradas. {len(fotos)} archivos.")
        generar_indice(fotos_dir, fotos)
        escribir_fecha_inspeccion(encargo, fotos)
        return 0

    # Renombrado en dos pasos para evitar colisiones
    tmp_pref = "_tmp_"
    for i, p in enumerate(pendientes, 1):
        tmp = p.parent / f"{tmp_pref}{i:04d}{p.suffix.lower()}"
        p.rename(tmp)

    # Reordenar todas (las ya-OK + las nuevas) por fecha EXIF/mtime
    todas = listar_fotos(fotos_dir)
    for i, p in enumerate(todas, 1):
        nuevo = p.parent / f"P{i:03d}{p.suffix.lower()}"
        if p != nuevo:
            p.rename(nuevo)

    final = listar_fotos(fotos_dir)
    print(f"OK — renombradas {len(final)} fotos como P001..P{len(final):03d}")

    fuentes = [fecha_orden(p)[1] for p in final]
    n_exif = fuentes.count("exif")
    if n_exif == len(final):
        print(f"Todas con EXIF: orden cronológico fiable.")
    elif n_exif == 0:
        print(f"AVISO: ninguna foto tiene EXIF; orden por fecha de archivo (poco fiable).")
    else:
        print(f"AVISO: {len(final) - n_exif}/{len(final)} fotos sin EXIF, ordenadas por mtime.")

    generar_indice(fotos_dir, final)
    escribir_fecha_inspeccion(encargo, final)
    xlsx = generar_excel_clasificacion(encargo)
    if xlsx.name:
        print(f"Excel de clasificación: {xlsx}")
        print(f"  → ábrelo, rellena Deficiencia y Descripción de cada foto, guarda, y ejecuta --aplicar-excel")
    return 0


def escribir_fecha_inspeccion(encargo: Path, fotos: list[Path]) -> None:
    """Si todas (o la mayoría) tienen EXIF, escribe el rango de fechas detectado."""
    fechas_exif = [fecha_orden(p)[0] for p in fotos if fecha_orden(p)[1] == "exif"]
    if not fechas_exif:
        return
    fmin = min(fechas_exif).date()
    fmax = max(fechas_exif).date()
    txt = encargo / "04_inspeccion" / "fecha_inspeccion.txt"
    if txt.exists() and txt.read_text(encoding="utf-8").strip():
        return  # respeta lo que Javier ya haya puesto
    if fmin == fmax:
        contenido = f"{fmin.isoformat()}\n"
    else:
        contenido = f"{fmin.isoformat()} a {fmax.isoformat()}\n"
    txt.write_text(contenido, encoding="utf-8")
    print(f"Escrita fecha de inspección detectada: {contenido.strip()}")


def generar_indice(carpeta: Path, fotos: list[Path]) -> None:
    md = ["# Índice de fotos\n", f"Total: **{len(fotos)}**\n"]
    md.append("| # | Archivo | Tamaño KB | Fecha tomada | Fuente |")
    md.append("|---|--------|----------:|--------------|--------|")
    for p in fotos:
        kb = round(p.stat().st_size / 1024)
        ts, src = fecha_orden(p)
        md.append(f"| {p.stem[1:]} | `{p.name}` | {kb} | {ts.strftime('%Y-%m-%d %H:%M')} | {src} |")
    (carpeta / "INDICE.md").write_text("\n".join(md) + "\n", encoding="utf-8")


def parse_rango(s: str) -> tuple[int, int]:
    if "-" in s:
        a, b = s.split("-", 1)
        return int(a), int(b)
    return int(s), int(s)


def parse_map(m: str) -> tuple[range, str, str]:
    """'1-12:DEF01:fachada_norte' -> (range(1,13), 'DEF01', 'fachada_norte')"""
    parts = m.split(":")
    if len(parts) != 3:
        raise ValueError(f"--map mal formado: {m!r}. Esperado 'A-B:DEFNN:slug'")
    rango_s, defcod, slug = parts
    a, b = parse_rango(rango_s)
    return range(a, b + 1), defcod, slug


def fase_clasificar(encargo: Path, mapeos: list[str]) -> int:
    fotos_dir = encargo / "04_inspeccion" / "fotos"
    dest_dir = encargo / "04_inspeccion" / "fotos_renombradas"
    dest_dir.mkdir(parents=True, exist_ok=True)

    fotos = listar_fotos(fotos_dir)
    by_num: dict[int, Path] = {}
    pat = re.compile(r"^P(\d{3,4})\.")
    for p in fotos:
        m = pat.match(p.name)
        if m:
            by_num[int(m.group(1))] = p

    if not by_num:
        print(f"ERROR: no se encontraron fotos con patrón Pxxx en {fotos_dir}", file=sys.stderr)
        print("       Ejecuta primero --renombrar", file=sys.stderr)
        return 1

    asignaciones: dict[int, tuple[str, str]] = {}
    for m in mapeos:
        rng, defcod, slug = parse_map(m)
        for n in rng:
            if n in asignaciones:
                print(f"WARN: foto {n} ya estaba asignada a {asignaciones[n]}, se sobrescribe", file=sys.stderr)
            asignaciones[n] = (defcod, slug)

    movidas = 0
    no_mapeadas = []
    for num, src in by_num.items():
        if num not in asignaciones:
            no_mapeadas.append(num)
            continue
        defcod, slug = asignaciones[num]
        dst = dest_dir / f"P{num:03d}_{defcod}_{slug}{src.suffix.lower()}"
        shutil.move(str(src), str(dst))
        movidas += 1

    print(f"OK — movidas {movidas} fotos a fotos_renombradas/")
    if no_mapeadas:
        print(f"AVISO: {len(no_mapeadas)} fotos sin mapeo, quedan en fotos/: {no_mapeadas[:20]}{'...' if len(no_mapeadas)>20 else ''}")
    return 0


def main(argv: list[str]) -> int:
    p = argparse.ArgumentParser()
    p.add_argument("--encargo", required=True, help="Ruta a la carpeta del encargo")
    g = p.add_mutually_exclusive_group(required=True)
    g.add_argument("--renombrar", action="store_true", help="Fase 1: copia desde --source y renombra; genera Excel de clasificación")
    g.add_argument("--aplicar-excel", action="store_true", dest="aplicar_excel",
                   help="Fase 2 RECOMENDADA: aplica clasificacion_fotos.xlsx y mueve a fotos_renombradas/")
    g.add_argument("--regenerar-excel", action="store_true", dest="regenerar_excel",
                   help="Regenera clasificacion_fotos.xlsx (mantiene los datos editados si se puede)")
    g.add_argument("--clasificar", action="store_true", help="Fase 2 LEGACY: aplica --map manualmente")
    p.add_argument("--source", default=None, help="Carpeta origen de Drive (sólo con --renombrar)")
    p.add_argument("--map", action="append", default=[], help='Mapeo rango:DEF:slug, ej. "1-12:DEF01:fachada_norte". Repetible.')
    args = p.parse_args(argv)

    encargo = Path(args.encargo)
    if not encargo.exists():
        print(f"ERROR: no existe {encargo}", file=sys.stderr)
        return 1

    if args.renombrar:
        source = Path(args.source) if args.source else None
        return fase_renombrar(encargo, source)
    if args.aplicar_excel:
        return fase_aplicar_excel(encargo)
    if args.regenerar_excel:
        xlsx = generar_excel_clasificacion(encargo)
        if xlsx.name:
            print(f"OK — Excel regenerado: {xlsx}")
            return 0
        print("ERROR: no se pudo generar el Excel", file=sys.stderr)
        return 1
    if args.clasificar:
        if not args.map:
            print("ERROR: --clasificar requiere al menos un --map", file=sys.stderr)
            return 1
        return fase_clasificar(encargo, args.map)
    return 1


if __name__ == "__main__":
    sys.exit(main(sys.argv[1:]))
